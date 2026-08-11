"""用户批量导入/导出（Excel，openpyxl）。"""
import io
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# 模板表头（顺序即列顺序）
HEADERS = [
    "用户名",
    "角色",
    "姓名/团队名称",
    "学工号",
    "邮箱",
    "初始密码（留空自动生成）",
    "成员姓名(JSON数组)",
    "成员学工号(JSON数组)",
]

ROLE_ALIAS = {
    "team": "team",
    "提交人": "team",
    "提交人（团队）": "team",
    "staff": "staff",
    "审核员": "staff",
    "admin": "admin",
    "管理员": "admin",
    "super_admin": "super_admin",
    "超管": "super_admin",
    "超级管理员": "super_admin",
}

# 导出账密表头
EXPORT_HEADERS = [
    "用户名", "角色", "姓名/团队名称", "学工号", "邮箱", "密码",
    "成员姓名", "成员学工号", "当前待办数", "累计完成数", "是否锁定",
]


def _to_bool_locked(locked_until) -> str:
    return "是" if locked_until else "否"


def _export_password(u) -> str:
    """仅导出系统设置的初始/重置密码（用户尚未自行修改过，password_changed_at 为空）。

    用户一旦自行修改密码（password_changed_at 非空），明文不再暴露，一律屏蔽，
    规避"导出账密"导致已修改密码泄露的风险。
    """
    if u.password_changed_at is None:
        return u.plain_password or ""
    return "已修改（不显示）"


def build_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用户数据"
    ws.append(HEADERS)
    ws.append([
        "team_example", "team", "示例团队", "", "team@example.com", "",
        '["张三","李四"]', '["2024001","2024002"]',
    ])
    # 角色下拉
    dv = DataValidation(
        type="list",
        formula1='"team,staff,admin,super_admin"',
        allow_blank=False,
        showErrorMessage=True,
        errorTitle="角色取值错误",
        error="角色只能填：team / staff / admin / super_admin",
    )
    ws.add_data_validation(dv)
    dv.add(f"B2:B1001")

    # 表头样式
    head_fill = PatternFill("solid", fgColor="4472C4")
    head_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center")
    for col, width in zip(HEADERS, (18, 14, 16, 12, 24, 22, 30, 30)):
        ws.column_dimensions[get_column_letter(HEADERS.index(col) + 1)].width = width
    ws.freeze_panes = "A2"

    sheet = wb.create_sheet("填写说明")
    tips = [
        "一、每行一个用户，请勿修改第一行表头。",
        "二、角色取值：team（提交人/团队）、staff（审核员）、admin（管理员）、super_admin（超级管理员）。",
        "三、审核员/管理员必须填写真实姓名与学工号，且学工号不可与其他审核员/管理员重复。",
        "四、提交人（team）必须填写成员姓名与成员学工号，两组JSON数组一一对应，且各不超过12人。",
        "五、初始密码留空则由系统自动生成8位临时密码（含大小写字母+数字+特殊字符），用户首次登录将强制改密。",
        "六、成员姓名示例：[\"张三\",\"李四\"]；也可用中文顿号分隔：张三、李四。",
        "七、导入前请删除示例行。",
    ]
    for t in tips:
        sheet.append([t])
    sheet.column_dimensions["A"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_cell(value):
    if value is None:
        return ""
    return str(value).strip()


def _parse_json_or_split(text: str) -> list[str]:
    text = (text or "").strip()
    if not text:
        return []
    if text.startswith("["):
        try:
            v = json.loads(text)
            if isinstance(v, list):
                return [str(x).strip() for x in v if str(x).strip()]
        except (json.JSONDecodeError, TypeError):
            pass
    parts = [p.strip() for p in text.replace("，", "、").replace(",", "、").split("、") if p.strip()]
    return parts


def parse_import_rows(content: bytes) -> list[dict]:
    """解析上传的 Excel：返回 [{row, username, data(原始字典)}, ...]，跳过全空行。"""
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    result = []
    for idx, values in enumerate(rows[1:], start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        record = {"row": idx}
        for pos, key in enumerate(HEADERS):
            record[key] = _parse_cell(values[pos]) if pos < len(values) else ""
        result.append(record)
    return result


def export_users(users) -> bytes:
    """导出全部用户：username/role/real_name/student_id/email/密码(仅初始密码，已改密者屏蔽)/成员信息/负载/锁定。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "账号密码"
    ws.append(EXPORT_HEADERS)
    for u in users:
        ws.append([
            u.username,
            u.role,
            u.real_name or "",
            u.student_id or "",
            u.email or "",
            _export_password(u),
            json.dumps(u.member_names or [], ensure_ascii=False),
            json.dumps(u.member_student_ids or [], ensure_ascii=False),
            u.current_pending_count,
            u.total_completed_count,
            _to_bool_locked(u.locked_until),
        ])
    head_fill = PatternFill("solid", fgColor="C00000")
    head_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = head_fill
        cell.font = head_font
    for i, w in enumerate((16, 14, 16, 12, 24, 20, 30, 30, 12, 12, 10), start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
