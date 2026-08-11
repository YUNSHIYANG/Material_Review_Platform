"""API 层冒烟测试：登录/锁定/强制改密/上传提交全链路。"""
import io

from fastapi.testclient import TestClient

from app.models import Submission

from conftest import make_submission, make_user


def _login(client: TestClient, username: str, password: str):
    return client.post("/api/auth/login", json={"username": username, "password": password})


def test_login_and_me(client, db):
    u = make_user(db, "staff", username="reviewer1", real_name="张三", student_id="S1", password_changed=True)
    db.commit()
    resp = _login(client, "reviewer1", "Passw0rd!x")
    assert resp.status_code == 200
    data = resp.json()
    assert data["role"] == "staff"
    assert data["need_password_change"] is False
    headers = {"Authorization": f"Bearer {data['access_token']}"}
    me = client.get("/api/auth/me", headers=headers)
    assert me.status_code == 200
    assert me.json()["username"] == "reviewer1"


def test_login_forced_password_change(client, db):
    u = make_user(db, "team", username="team1", real_name="财务部", password_changed=False)
    db.commit()
    resp = _login(client, "team1", "Passw0rd!x")
    assert resp.status_code == 200
    assert resp.json()["need_password_change"] is True
    headers = {"Authorization": f"Bearer {resp.json()['access_token']}"}
    # 未改密时访问受保护接口被中间件拦截
    r = client.get("/api/team/submissions", headers=headers)
    assert r.status_code == 403
    assert r.json()["detail"] == "NEED_PASSWORD_CHANGE"
    # 修改密码后放行
    r = client.post("/api/auth/change-password", headers=headers,
                    json={"old_password": "Passw0rd!x", "new_password": "NewPassw0rd!x"})
    assert r.status_code == 200
    r = client.get("/api/team/submissions", headers=headers)
    assert r.status_code == 200


def test_login_lockout_after_five_failures(client, db):
    u = make_user(db, "staff", username="lockuser", real_name="李四", student_id="S2", password_changed=True)
    db.commit()
    # 前4次失败：仅计数，返回 400
    for i in range(4):
        resp = _login(client, "lockuser", "WrongPass1!")
        assert resp.status_code == 400
    # 第5次失败：触发锁定，返回 423
    resp = _login(client, "lockuser", "WrongPass1!")
    assert resp.status_code == 423
    # 锁定期间即使密码正确也返回 423
    resp = _login(client, "lockuser", "Passw0rd!x")
    assert resp.status_code == 423


def test_team_submit_full_flow(client, db):
    team = make_user(db, "team", username="team2", real_name="材料学院",
                     member_names=["王五"], member_student_ids=["2024001"], password_changed=True)
    s1 = make_user(db, "staff", username="staff1", real_name="甲", student_id="A1", password_changed=True)
    s2 = make_user(db, "staff", username="staff2", real_name="乙", student_id="A2", password_changed=True)
    ad1 = make_user(db, "admin", username="admin1", real_name="管理员甲", student_id="AD1", password_changed=True)
    ad2 = make_user(db, "admin", username="admin2", real_name="管理员乙", student_id="AD2", password_changed=True)
    db.commit()

    # 团队登录并上传
    resp = _login(client, "team2", "Passw0rd!x")
    token = resp.json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    files = {"file": ("2024年报销单.zip", io.BytesIO(b"PK\x03\x04 fake zip content"), "application/zip")}
    resp = client.post("/api/team/submissions", headers=headers, data={"remark": "请尽快审核"}, files=files)
    assert resp.status_code == 200, resp.text
    sub_id = resp.json()["submission_id"]

    sub = db.get(Submission, sub_id)
    assert sub.status == "first_reviewing"
    assert len(sub.staff_reviews) == 2

    # 团队看板
    dashboard = client.get("/api/team/submissions", headers=headers)
    assert dashboard.status_code == 200
    assert dashboard.json()[0]["user_status"] == "审核中"

    # 审核员1登录提交意见
    r1 = _login(client, "staff1", "Passw0rd!x")
    h1 = {"Authorization": f"Bearer {r1.json()['access_token']}"}
    r = client.post(f"/api/staff/submissions/{sub_id}/review", headers=h1,
                    json={"result": True, "comment": "材料齐全", "version": sub.version})
    assert r.status_code == 200, r.text
    db.refresh(sub)
    assert sub.status == "first_reviewing"

    # 审核员2提交 → 进入终审
    r2 = _login(client, "staff2", "Passw0rd!x")
    h2 = {"Authorization": f"Bearer {r2.json()['access_token']}"}
    r = client.post(f"/api/staff/submissions/{sub_id}/review", headers=h2,
                    json={"result": True, "comment": "", "version": sub.version})
    assert r.status_code == 200, r.text
    db.refresh(sub)
    assert sub.status == "admin_reviewing"
    admin_id = sub.assigned_admin_id

    # 管理员裁定通过
    admin_name = "admin1" if admin_id == ad1.id else "admin2"
    ra = _login(client, admin_name, "Passw0rd!x")
    ha = {"Authorization": f"Bearer {ra.json()['access_token']}"}
    r = client.post(f"/api/admin/submissions/{sub_id}/review", headers=ha,
                    json={"final_result": True, "admin_comment": "同意", "version": sub.version})
    assert r.status_code == 200, r.text
    db.refresh(sub)
    assert sub.status == "passed"

    # 团队端看到已通过
    dashboard = client.get("/api/team/submissions", headers=headers)
    assert dashboard.json()[0]["user_status"] == "已通过"


def test_withdraw_decrements_reviewer_pending(client, db):
    """回归：团队提交后立即撤回，审核员待办应正确扣减；第2轮重新提交后重新累加。"""
    team = make_user(db, "team", username="teamw", real_name="团队w",
                     member_names=["王五"], member_student_ids=["2024001"], password_changed=True)
    s1 = make_user(db, "staff", username="sw1", real_name="甲", student_id="W1", password_changed=True)
    s2 = make_user(db, "staff", username="sw2", real_name="乙", student_id="W2", password_changed=True)
    db.commit()

    t = _login(client, "teamw", "Passw0rd!x")
    th = {"Authorization": f"Bearer {t.json()['access_token']}"}
    files = {"file": ("a.zip", io.BytesIO(b"PK\x03\x04 x"), "application/zip")}
    r = client.post("/api/team/submissions", headers=th, data={"remark": "x"}, files=files)
    assert r.status_code == 200, r.text
    sub_id = r.json()["submission_id"]
    db.refresh(s1)
    db.refresh(s2)
    assert (s1.current_pending_count, s2.current_pending_count) == (1, 1)

    # 立即撤回 → 待办归零
    sub = db.get(Submission, sub_id)
    r = client.post(f"/api/team/submissions/{sub_id}/withdraw", headers=th, json={"version": sub.version})
    assert r.status_code == 200, r.text
    db.refresh(s1)
    db.refresh(s2)
    assert (s1.current_pending_count, s2.current_pending_count) == (0, 0)

    # 第2轮重新提交 → 待办重新累加（测试员此前看到的"1"即来自新一轮分配）
    r = client.post("/api/team/submissions", headers=th, data={"remark": "x2"}, files=files)
    assert r.status_code == 200, r.text
    db.refresh(s1)
    db.refresh(s2)
    assert (s1.current_pending_count, s2.current_pending_count) == (1, 1)


def test_load_monitor_includes_staff_and_admin(client, db):
    """回归：负载监控须同时返回全部审核员（staff）与管理员（admin）。"""
    make_user(db, "super_admin", username="bossL", real_name="超管", password_changed=True)
    st = make_user(db, "staff", username="sl1", real_name="审核员L", student_id="SL1", password_changed=True)
    ad = make_user(db, "admin", username="al1", real_name="管理员L", student_id="AL1", password_changed=True)
    db.commit()

    resp = _login(client, "bossL", "Passw0rd!x")
    h = {"Authorization": f"Bearer {resp.json()['access_token']}"}
    r = client.get("/api/super/load", headers=h)
    assert r.status_code == 200, r.text
    rows = r.json()
    ids = {row["id"]: row["role"] for row in rows}
    assert ids.get(st.id) == "staff"
    assert ids.get(ad.id) == "admin"


def test_super_sensitive_ops_no_422(client, db):
    """回归：超管敏感操作（需密码二次确认）不再因 request 参数解析错误返回 422。"""
    make_user(db, "super_admin", username="boss", real_name="超管", password_changed=True)
    target = make_user(db, "staff", username="r1", real_name="审核员1", student_id="S1", password_changed=True)
    team = make_user(db, "team", username="teamx", real_name="团队x",
                     member_names=["王五"], member_student_ids=["2024001"], password_changed=True)
    db.commit()
    sub = make_submission(db, team, version=3,
                          first_review_skip_reason="insufficient_staff", assigned_admin_id=None)
    sub.status = "pending_admin_intervention"
    db.commit()

    resp = _login(client, "boss", "Passw0rd!x")
    h = {"Authorization": f"Bearer {resp.json()['access_token']}"}
    r = client.post(f"/api/super/users/{target.id}/reset-password", headers=h,
                    json={"password": "Passw0rd!x"})
    assert r.status_code == 200, r.text
    r = client.post(f"/api/super/submissions/{sub.id}/intervene", headers=h,
                    json={"action": "force_pass", "version": 3, "password": "Passw0rd!x", "comment": "测试"})
    assert r.status_code == 200, r.text
    r = client.put("/api/super/config", headers=h,
                   json={"timeout_hours": 72, "password": "Passw0rd!x"})
    assert r.status_code == 200, r.text


def test_user_import_export_xlsx(client, db):
    """回归：Excel 模板下载、批量导入用户、导出账密（含明文密码）。"""
    import openpyxl

    from app.models import User

    make_user(db, "super_admin", username="boss2", real_name="超管2", password_changed=True)
    db.commit()
    resp = _login(client, "boss2", "Passw0rd!x")
    h = {"Authorization": f"Bearer {resp.json()['access_token']}"}

    # 模板下载
    r = client.get("/api/super/users/template", headers=h)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # 构造导入文件：1 个审核员 + 1 个团队
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["用户名", "角色", "姓名/团队名称", "学工号", "邮箱",
               "初始密码（留空自动生成）", "成员姓名(JSON数组)", "成员学工号(JSON数组)"])
    ws.append(["staff_imp", "staff", "导入审核员", "S999", "staff_imp@example.com", "Staff@1234", "", ""])
    ws.append(["team_imp", "team", "导入团队", "", "team_imp@example.com", "", '["张三","李四"]', '["2024001","2024002"]'])
    ws.append(["bad_row", "unknown_role", "错误角色", "", "bad@example.com", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)

    r = client.post("/api/super/users/import", headers=h,
                    files={"file": ("import.xlsx", buf.getvalue(),
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["created_count"] == 2
    assert len(data["errors"]) == 1  # bad_row 角色无效

    staff_imp = db.query(User).filter(User.username == "staff_imp").first()
    assert staff_imp is not None
    assert staff_imp.plain_password == "Staff@1234"
    team_imp = db.query(User).filter(User.username == "team_imp").first()
    assert team_imp.member_names == ["张三", "李四"]
    assert team_imp.team_id == team_imp.id

    # 导出账密：包含明文密码
    r = client.get("/api/super/users/export", headers=h)
    assert r.status_code == 200, r.text
    out = openpyxl.load_workbook(io.BytesIO(r.content))
    rows = list(out.active.iter_rows(values_only=True))
    headers = rows[0]
    pwd_idx = headers.index("密码")
    pwd_by_username = {row[0]: row[pwd_idx] for row in rows[1:]}
    assert pwd_by_username.get("staff_imp") == "Staff@1234"

    # 用户自行修改密码后，导出必须屏蔽明文（仅初始密码可导出）
    from app.utils import utcnow

    staff_imp.password_changed_at = utcnow()
    db.commit()
    r = client.get("/api/super/users/export", headers=h)
    assert r.status_code == 200, r.text
    rows2 = list(openpyxl.load_workbook(io.BytesIO(r.content)).active.iter_rows(values_only=True))
    pwd2 = {row[0]: row[rows2[0].index("密码")] for row in rows2[1:]}
    assert pwd2["staff_imp"] != "Staff@1234"


def _flow_to_passed(client, db, team_username="teamflow", suffix=""):
    """完整流程到『已通过』：上传 → 双初审 → 管理员通过。返回 (team_headers, sub_id)。"""
    team = make_user(db, "team", username=f"{team_username}{suffix}", real_name=f"团队{suffix}",
                     member_names=["王五"], member_student_ids=["2024{suffix}"], password_changed=True)
    s1 = make_user(db, "staff", username=f"sf1{suffix}", real_name="甲", student_id=f"F1{suffix}", password_changed=True)
    s2 = make_user(db, "staff", username=f"sf2{suffix}", real_name="乙", student_id=f"F2{suffix}", password_changed=True)
    ad1 = make_user(db, "admin", username=f"ad1{suffix}", real_name="管理员甲", student_id=f"D1{suffix}", password_changed=True)
    ad2 = make_user(db, "admin", username=f"ad2{suffix}", real_name="管理员乙", student_id=f"D2{suffix}", password_changed=True)
    db.commit()

    t = _login(client, f"{team_username}{suffix}", "Passw0rd!x")
    th = {"Authorization": f"Bearer {t.json()['access_token']}"}
    files = {"file": ("报销单.zip", io.BytesIO(b"PK\x03\x04 flow content"), "application/zip")}
    r = client.post("/api/team/submissions", headers=th, data={"remark": "x"}, files=files)
    assert r.status_code == 200, r.text
    sub_id = r.json()["submission_id"]
    sub = db.get(Submission, sub_id)

    for name in (f"sf1{suffix}", f"sf2{suffix}"):
        rr = _login(client, name, "Passw0rd!x")
        hh = {"Authorization": f"Bearer {rr.json()['access_token']}"}
        r = client.post(f"/api/staff/submissions/{sub_id}/review", headers=hh,
                        json={"result": True, "comment": "ok", "version": sub.version})
        assert r.status_code == 200, r.text
        db.refresh(sub)

    admin_name = f"ad1{suffix}" if sub.assigned_admin_id == ad1.id else f"ad2{suffix}"
    ra = _login(client, admin_name, "Passw0rd!x")
    ha = {"Authorization": f"Bearer {ra.json()['access_token']}"}
    r = client.post(f"/api/admin/submissions/{sub_id}/review", headers=ha,
                    json={"final_result": True, "admin_comment": "同意", "version": sub.version})
    assert r.status_code == 200, r.text
    db.refresh(sub)
    assert sub.status == "passed"
    return th, sub_id


def test_withdraw_after_passed_allows_resubmit(client, db):
    """已通过的工单允许团队撤回，撤回后可重新提交并重走流程。"""
    th, sub_id = _flow_to_passed(client, db, suffix="W")
    sub = db.get(Submission, sub_id)

    # 团队端可撤回
    detail = client.get(f"/api/team/submissions/{sub_id}", headers=th)
    assert detail.status_code == 200
    assert detail.json()["can_withdraw"] is True

    r = client.post(f"/api/team/submissions/{sub_id}/withdraw", headers=th, json={"version": sub.version})
    assert r.status_code == 200, r.text
    db.refresh(sub)
    assert sub.status == "withdrawn"

    # 撤回后重新提交 → 新一轮
    files = {"file": ("报销单2.zip", io.BytesIO(b"PK\x03\x04 resubmit"), "application/zip")}
    r = client.post("/api/team/submissions", headers=th, data={"remark": "重提"}, files=files)
    assert r.status_code == 200, r.text
    new_id = r.json()["submission_id"]
    new_sub = db.get(Submission, new_id)
    assert new_sub.submit_round == 2
    assert new_sub.status == "first_reviewing"


def test_super_return_passed_submission(client, db):
    """超管打回已通过材料：附意见、状态置为 returned、团队收到打回意见并可重新提交。"""
    make_user(db, "super_admin", username="bossRet", real_name="超管", password_changed=True)
    db.commit()
    th, sub_id = _flow_to_passed(client, db, suffix="R")
    sub = db.get(Submission, sub_id)

    rb = _login(client, "bossRet", "Passw0rd!x")
    h = {"Authorization": f"Bearer {rb.json()['access_token']}"}

    # 未填意见 → 400
    r = client.post(f"/api/super/submissions/{sub_id}/return", headers=h,
                    json={"password": "Passw0rd!x", "version": sub.version, "comment": ""})
    assert r.status_code == 400, r.text

    r = client.post(f"/api/super/submissions/{sub_id}/return", headers=h,
                    json={"password": "Passw0rd!x", "version": sub.version, "comment": "发票金额不符，请修改"})
    assert r.status_code == 200, r.text
    db.refresh(sub)
    assert sub.status == "returned"
    assert sub.return_comment == "发票金额不符，请修改"
    assert sub.returned_at is not None

    # 团队端可见打回意见，且不可撤回
    detail = client.get(f"/api/team/submissions/{sub_id}", headers=th)
    assert detail.status_code == 200
    assert "发票金额不符" in detail.json()["returned_note"]
    assert detail.json()["can_withdraw"] is False

    # 团队重新提交 → 新一轮（父工单追溯打回记录）
    files = {"file": ("报销单3.zip", io.BytesIO(b"PK\x03\x04 resubmit2"), "application/zip")}
    r = client.post("/api/team/submissions", headers=th, data={"remark": "重提"}, files=files)
    assert r.status_code == 200, r.text
    new_sub = db.get(Submission, r.json()["submission_id"])
    assert new_sub.submit_round == 2
    assert new_sub.parent_submission_id is not None


def test_bulk_download_materials_zip(client, db):
    """超管一键下载材料：zip 内目录与后台存储端命名一致（团队名_id）、仅含最新 version、支持状态筛选。"""
    import os
    import zipfile

    from app.config import get_settings
    from app.utils import sanitize_dirname

    make_user(db, "super_admin", username="bossDL", real_name="超管", password_changed=True)
    team_a = make_user(db, "team", username="teamA", real_name="团队A 材料/一",
                       member_names=["王五"], member_student_ids=["A1"], password_changed=True)
    db.commit()
    settings = get_settings()
    folder = f"{sanitize_dirname(team_a.real_name, fallback='team_%s' % team_a.id)}_{team_a.id}"
    os.makedirs(os.path.join(settings.upload_dir, folder), exist_ok=True)
    f1 = os.path.join(folder, "round1_materials.zip")
    f2 = os.path.join(folder, "round2_materials.zip")
    for rel, content in ((f1, b"ROUND1"), (f2, b"ROUND2")):
        with open(os.path.join(settings.upload_dir, rel), "wb") as fh:
            fh.write(content)
    make_submission(db, team_a, submit_round=1, status="passed", file_path=f1, file_stored_name="round1_materials.zip")
    make_submission(db, team_a, submit_round=2, status="passed", file_path=f2, file_stored_name="round2_materials.zip")
    db.commit()

    rb = _login(client, "bossDL", "Passw0rd!x")
    h = {"Authorization": f"Bearer {rb.json()['access_token']}"}
    r = client.get("/api/super/submissions/download?status=passed", headers=h)
    assert r.status_code == 200, r.text
    assert r.headers["content-type"] == "application/zip"

    zf = zipfile.ZipFile(io.BytesIO(r.content))
    names = zf.namelist()
    # 每支队伍仅最新 version（round2）
    entries = [n for n in names if n.startswith(folder)]
    assert entries == [f"{folder}/round2_materials.zip"]
    assert zf.read(f"{folder}/round2_materials.zip") == b"ROUND2"
